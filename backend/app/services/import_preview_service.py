import csv
import re
import unicodedata
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import PurePath
from typing import Any

from app.schemas.import_data import SUPPORTED_IMPORT_EXTENSIONS


DEFAULT_PREVIEW_ROW_LIMIT = 50
SUPPORTED_CSV_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030")
MAX_IDENTIFIER_LENGTH = 63


class ImportPreviewError(ValueError):
    pass


class UnsupportedImportFileError(ImportPreviewError):
    pass


class EmptyImportFileError(ImportPreviewError):
    pass


class ImportPreviewService:
    def preview_file(
        self,
        *,
        filename: str,
        content: bytes,
        content_type: str = "",
        preview_limit: int | None = DEFAULT_PREVIEW_ROW_LIMIT,
    ) -> dict[str, Any]:
        clean_filename = PurePath(filename or "").name
        extension = extension_from_filename(clean_filename)
        if not clean_filename:
            raise ImportPreviewError("上传文件名不能为空。")
        if extension not in SUPPORTED_IMPORT_EXTENSIONS:
            raise UnsupportedImportFileError("只支持上传 CSV 或 XLSX 文件。")
        if not content:
            raise EmptyImportFileError("上传文件不能为空。")

        if extension == "csv":
            rows, metadata = self._read_csv(content)
        else:
            rows, metadata = self._read_xlsx(content)

        parsed = self._build_preview(rows=rows, preview_limit=preview_limit)
        return {
            "filename": clean_filename,
            "extension": extension,
            "content_type": content_type or "application/octet-stream",
            "sheet_name": metadata.get("sheet_name"),
            "encoding": metadata.get("encoding"),
            **parsed,
        }

    def _read_csv(self, content: bytes) -> tuple[list[list[Any]], dict[str, Any]]:
        text = None
        encoding_used = None
        for encoding in SUPPORTED_CSV_ENCODINGS:
            try:
                text = content.decode(encoding)
                encoding_used = encoding
                break
            except UnicodeDecodeError:
                continue

        if text is None or encoding_used is None:
            raise ImportPreviewError("CSV 文件编码不支持，请使用 UTF-8 或 GB18030 编码。")

        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(StringIO(text), dialect)
        return [list(row) for row in reader], {"encoding": encoding_used}

    def _read_xlsx(self, content: bytes) -> tuple[list[list[Any]], dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportPreviewError("后端缺少 XLSX 解析依赖 openpyxl。") from exc

        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ImportPreviewError("XLSX 文件无法读取，请确认文件没有损坏。") from exc

        worksheet = workbook.worksheets[0]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        return rows, {"sheet_name": worksheet.title}

    def _build_preview(self, *, rows: list[list[Any]], preview_limit: int | None) -> dict[str, Any]:
        non_empty_rows = [row for row in rows if not _is_blank_row(row)]
        removed_empty_rows = len(rows) - len(non_empty_rows)
        if not non_empty_rows:
            raise EmptyImportFileError("文件中没有可预览的数据。")

        rows_without_empty_columns, removed_empty_columns = _remove_empty_columns(non_empty_rows)
        if not rows_without_empty_columns or not rows_without_empty_columns[0]:
            raise EmptyImportFileError("文件中没有有效字段。")

        header_index = _find_header_index(rows_without_empty_columns)
        header_row = rows_without_empty_columns[header_index]
        data_rows = [row for row in rows_without_empty_columns[header_index + 1 :] if not _is_blank_row(row)]
        column_count = len(header_row)
        padded_data_rows = [_pad_row(row, column_count) for row in data_rows]

        names, name_warnings = _normalize_column_names(header_row)
        columns = []
        for index, name in enumerate(names):
            values = [row[index] for row in padded_data_rows]
            columns.append(
                {
                    "name": name,
                    "original_name": _stringify_header(header_row[index]),
                    "type": _infer_column_type(values),
                    "nullable": any(_is_blank_value(value) for value in values),
                    "sample_values": _sample_values(values),
                }
            )

        preview_rows = []
        preview_data_rows = padded_data_rows if preview_limit is None else padded_data_rows[:preview_limit]
        for row in preview_data_rows:
            preview_rows.append({column["name"]: _jsonable_value(row[index]) for index, column in enumerate(columns)})

        warnings = []
        if header_index > 0:
            warnings.append(f"已跳过表头前的 {header_index} 行说明性内容。")
        if removed_empty_rows > 0:
            warnings.append(f"已清理 {removed_empty_rows} 个空行。")
        if removed_empty_columns > 0:
            warnings.append(f"已清理 {removed_empty_columns} 个空列。")
        warnings.extend(name_warnings)
        if preview_limit is not None and len(padded_data_rows) > preview_limit:
            warnings.append(f"当前只预览前 {preview_limit} 行，导入时会处理全部数据。")

        return {
            "columns": columns,
            "rows": preview_rows,
            "row_count": len(padded_data_rows),
            "column_count": len(columns),
            "preview_row_count": len(preview_rows),
            "warnings": warnings,
        }


def extension_from_filename(filename: str) -> str:
    suffix = PurePath(filename).suffix.lower()
    if not suffix:
        return ""
    return suffix.removeprefix(".")


def _remove_empty_columns(rows: list[list[Any]]) -> tuple[list[list[Any]], int]:
    max_columns = max(len(row) for row in rows)
    padded_rows = [_pad_row(row, max_columns) for row in rows]
    keep_indexes = [
        index
        for index in range(max_columns)
        if any(not _is_blank_value(row[index]) for row in padded_rows)
    ]
    cleaned = [[row[index] for index in keep_indexes] for row in padded_rows]
    return cleaned, max_columns - len(keep_indexes)


def _find_header_index(rows: list[list[Any]]) -> int:
    first_non_empty_index = 0
    for index, row in enumerate(rows):
        if not _is_blank_row(row):
            first_non_empty_index = index
            break

    for index, row in enumerate(rows[first_non_empty_index : first_non_empty_index + 10], start=first_non_empty_index):
        non_empty_count = sum(1 for value in row if not _is_blank_value(value))
        if non_empty_count >= 2:
            return index
    return first_non_empty_index


def _normalize_column_names(header_row: list[Any]) -> tuple[list[str], list[str]]:
    names = []
    warnings = []
    used_names: set[str] = set()
    for index, value in enumerate(header_row):
        original_name = _stringify_header(value)
        normalized = _normalize_identifier(original_name)
        generated_from_empty = False
        if not normalized:
            normalized = f"column_{index + 1}"
            generated_from_empty = True
        elif normalized[0].isdigit():
            normalized = f"column_{normalized}"
        normalized = normalized[:MAX_IDENTIFIER_LENGTH]

        unique_name = _deduplicate_name(normalized, used_names)
        used_names.add(unique_name)
        names.append(unique_name)

        if not original_name:
            warnings.append(f"第 {index + 1} 列表头为空，已命名为 {unique_name}。")
        elif unique_name != normalized:
            warnings.append(f"字段名 {original_name} 重复，已重命名为 {unique_name}。")
        elif generated_from_empty:
            warnings.append(f"字段名 {original_name} 已规范化为 {unique_name}。")
        elif unique_name != original_name:
            warnings.append(f"字段名 {original_name} 已规范化为 {unique_name}。")

    return names, warnings


def _normalize_identifier(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    safe_name = re.sub(r"[^0-9A-Za-z]+", "_", ascii_value).strip("_").lower()
    return re.sub(r"_+", "_", safe_name)[:MAX_IDENTIFIER_LENGTH]


def _deduplicate_name(name: str, used_names: set[str]) -> str:
    if name not in used_names:
        return name

    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = f"{name[: MAX_IDENTIFIER_LENGTH - len(suffix)]}{suffix}"
        if candidate not in used_names:
            return candidate
        counter += 1


def _infer_column_type(values: list[Any]) -> str:
    non_empty_values = [value for value in values if not _is_blank_value(value)]
    if not non_empty_values:
        return "text"
    if all(_is_bool_value(value) for value in non_empty_values):
        return "boolean"
    if all(_is_integer_value(value) for value in non_empty_values):
        return "integer"
    if all(_is_numeric_value(value) for value in non_empty_values):
        return "numeric"
    if all(_is_date_value(value) for value in non_empty_values):
        return "date"
    if all(_is_timestamp_value(value) or _is_date_value(value) for value in non_empty_values):
        return "timestamp"
    return "text"


def _is_bool_value(value: Any) -> bool:
    if isinstance(value, bool):
        return True
    if isinstance(value, str):
        return value.strip().lower() in {"true", "false", "yes", "no", "y", "n"}
    return False


def _is_integer_value(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    if isinstance(value, str):
        return bool(re.fullmatch(r"[+-]?\d+", value.strip()))
    return False


def _is_numeric_value(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int | float):
        return True
    if isinstance(value, str):
        return bool(re.fullmatch(r"[+-]?(\d+(\.\d*)?|\.\d+)", value.strip()))
    return False


def _is_date_value(value: Any) -> bool:
    if isinstance(value, datetime):
        return False
    if isinstance(value, date):
        return True
    return _parse_date_string(value) is not None


def _is_timestamp_value(value: Any) -> bool:
    if isinstance(value, datetime):
        return True
    return _parse_datetime_string(value) is not None


def _parse_date_string(value: Any) -> date | None:
    if not isinstance(value, str):
        return None
    text = value.strip()
    for date_format in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def _parse_datetime_string(value: Any) -> datetime | None:
    if not isinstance(value, str):
        return None
    text = value.strip()
    for datetime_format in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
    ):
        try:
            return datetime.strptime(text, datetime_format)
        except ValueError:
            continue
    return None


def _sample_values(values: list[Any], limit: int = 3) -> list[Any]:
    samples = []
    seen = set()
    for value in values:
        if _is_blank_value(value):
            continue
        json_value = _jsonable_value(value)
        marker = repr(json_value)
        if marker in seen:
            continue
        seen.add(marker)
        samples.append(json_value)
        if len(samples) >= limit:
            break
    return samples


def _jsonable_value(value: Any) -> Any:
    if _is_blank_value(value):
        return None
    if isinstance(value, datetime | date):
        return value.isoformat()
    return value


def _stringify_header(value: Any) -> str:
    if _is_blank_value(value):
        return ""
    return str(value).strip()


def _is_blank_row(row: list[Any]) -> bool:
    return all(_is_blank_value(value) for value in row)


def _is_blank_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _pad_row(row: list[Any], length: int) -> list[Any]:
    return [*row, *([None] * (length - len(row)))]
